import os
import hashlib
import secrets
import tempfile
from datetime import datetime

import gradio as gr
import pandas as pd

from sqlalchemy import (
    create_engine,
    Column,
    Integer,
    String,
    Boolean,
    DateTime,
    ForeignKey,
    UniqueConstraint,
)

from sqlalchemy.orm import (
    declarative_base,
    sessionmaker,
)


# ==========================================================
# SCHOOL MARK ENTRY SYSTEM
# ADMIN + TEACHER + STUDENT + EXCEL VERSION
# ==========================================================


# ==========================================================
# DATABASE CONFIGURATION
# ==========================================================

DATABASE_URL = os.getenv(
    "DATABASE_URL",
    "",
).strip()


if DATABASE_URL.startswith("postgres://"):
    DATABASE_URL = (
        "postgresql://"
        + DATABASE_URL[11:]
    )


if DATABASE_URL:

    engine = create_engine(
        DATABASE_URL,
        pool_pre_ping=True,
    )

else:

    engine = create_engine(
        "sqlite:///school_marks.db",
        connect_args={
            "check_same_thread": False
        },
    )


SessionLocal = sessionmaker(
    bind=engine,
)


Base = declarative_base()


# ==========================================================
# DATABASE TABLES
# ==========================================================


class AcademicYear(Base):

    __tablename__ = "academic_years"

    id = Column(
        Integer,
        primary_key=True,
    )

    name = Column(
        String(50),
        unique=True,
        nullable=False,
    )

    active = Column(
        Boolean,
        default=True,
    )


class ClassSection(Base):

    __tablename__ = "classes"

    id = Column(
        Integer,
        primary_key=True,
    )

    name = Column(
        String(50),
        unique=True,
        nullable=False,
    )


class Student(Base):

    __tablename__ = "students"

    id = Column(
        Integer,
        primary_key=True,
    )

    admission_no = Column(
        String(50),
        unique=True,
        nullable=False,
    )

    roll_no = Column(
        String(50),
    )

    name = Column(
        String(200),
        nullable=False,
    )

    class_id = Column(
        Integer,
        ForeignKey("classes.id"),
        nullable=False,
    )

    active = Column(
        Boolean,
        default=True,
    )


class Teacher(Base):

    __tablename__ = "teachers"

    id = Column(
        Integer,
        primary_key=True,
    )

    username = Column(
        String(100),
        unique=True,
        nullable=False,
    )

    name = Column(
        String(200),
        nullable=False,
    )

    password_hash = Column(
        String(500),
        nullable=False,
    )

    role = Column(
        String(20),
        default="teacher",
    )

    active = Column(
        Boolean,
        default=True,
    )


class Subject(Base):

    __tablename__ = "subjects"

    id = Column(
        Integer,
        primary_key=True,
    )

    name = Column(
        String(200),
        nullable=False,
    )

    code = Column(
        String(50),
        unique=True,
        nullable=False,
    )

    theory = Column(
        Boolean,
        default=True,
    )

    practical = Column(
        Boolean,
        default=False,
    )

    internal = Column(
        Boolean,
        default=True,
    )

    theory_max = Column(
        Integer,
        default=0,
    )

    practical_max = Column(
        Integer,
        default=0,
    )

    internal_max = Column(
        Integer,
        default=0,
    )

    active = Column(
        Boolean,
        default=True,
    )


class ClassSubject(Base):

    __tablename__ = "class_subjects"

    id = Column(
        Integer,
        primary_key=True,
    )

    class_id = Column(
        Integer,
        ForeignKey("classes.id"),
        nullable=False,
    )

    subject_id = Column(
        Integer,
        ForeignKey("subjects.id"),
        nullable=False,
    )

    __table_args__ = (
        UniqueConstraint(
            "class_id",
            "subject_id",
            name="uq_class_subject",
        ),
    )


class Exam(Base):

    __tablename__ = "exams"

    id = Column(
        Integer,
        primary_key=True,
    )

    name = Column(
        String(100),
        unique=True,
        nullable=False,
    )

    active = Column(
        Boolean,
        default=True,
    )


class Mark(Base):

    __tablename__ = "marks"

    id = Column(
        Integer,
        primary_key=True,
    )

    academic_year_id = Column(
        Integer,
        ForeignKey("academic_years.id"),
        nullable=False,
    )

    exam_id = Column(
        Integer,
        ForeignKey("exams.id"),
        nullable=False,
    )

    student_id = Column(
        Integer,
        ForeignKey("students.id"),
        nullable=False,
    )

    subject_id = Column(
        Integer,
        ForeignKey("subjects.id"),
        nullable=False,
    )

    theory = Column(
        Integer,
        default=0,
    )

    practical = Column(
        Integer,
        default=0,
    )

    internal = Column(
        Integer,
        default=0,
    )

    total = Column(
        Integer,
        default=0,
    )

    updated_at = Column(
        DateTime,
        default=datetime.utcnow,
    )

    __table_args__ = (
        UniqueConstraint(
            "academic_year_id",
            "exam_id",
            "student_id",
            "subject_id",
            name="uq_student_exam_subject",
        ),
    )


# ==========================================================
# CREATE TABLES
# ==========================================================

Base.metadata.create_all(
    engine
)


# ==========================================================
# PASSWORD FUNCTIONS
# ==========================================================


def hash_password(password):

    salt = secrets.token_hex(16)

    digest = hashlib.pbkdf2_hmac(
        "sha256",
        str(password).encode("utf-8"),
        salt.encode("utf-8"),
        120000,
    ).hex()

    return (
        salt
        + "$"
        + digest
    )


def verify_password(
    password,
    stored,
):

    try:

        salt, digest = stored.split(
            "$",
            1,
        )

        check = hashlib.pbkdf2_hmac(
            "sha256",
            str(password).encode("utf-8"),
            salt.encode("utf-8"),
            120000,
        ).hex()

        return secrets.compare_digest(
            check,
            digest,
        )

    except Exception:

        return False


# ==========================================================
# BASIC HELPERS
# ==========================================================


def clean_name(value):

    return str(
        value or ""
    ).strip()


def parse_id(value):

    if value is None:
        return None

    text = str(value).strip()

    if not text:
        return None

    try:

        return int(
            text.split("|", 1)[0].strip()
        )

    except Exception:

        return None


def safe_int(
    value,
    default=0,
):

    try:

        if value in (
            None,
            "",
        ):

            return default

        return int(
            float(value)
        )

    except Exception:

        return default


def empty_dropdown():

    return gr.Dropdown(
        choices=[],
        value=None,
    )


# ==========================================================
# INITIAL DATABASE
# ==========================================================


def seed_database():

    db = SessionLocal()

    try:

        # --------------------------------------------------
        # Academic Year
        # --------------------------------------------------

        if not db.query(
            AcademicYear
        ).first():

            db.add(
                AcademicYear(
                    name="2026-27",
                    active=True,
                )
            )


        # --------------------------------------------------
        # Classes
        # --------------------------------------------------

        if not db.query(
            ClassSection
        ).first():

            db.add_all([

                ClassSection(
                    name="10-A"
                ),

                ClassSection(
                    name="10-B"
                ),

                ClassSection(
                    name="11-A"
                ),

                ClassSection(
                    name="11-B"
                ),

                ClassSection(
                    name="12-A"
                ),

                ClassSection(
                    name="12-B"
                ),

            ])


        # --------------------------------------------------
        # Exams
        # --------------------------------------------------

        if not db.query(
            Exam
        ).first():

            db.add_all([

                Exam(
                    name="Unit Test 1"
                ),

                Exam(
                    name="Quarterly"
                ),

                Exam(
                    name="Half-Yearly"
                ),

                Exam(
                    name="Annual"
                ),

            ])


        # --------------------------------------------------
        # Subjects
        # --------------------------------------------------

        if not db.query(
            Subject
        ).first():

            db.add_all([

                Subject(
                    name="Tamil",
                    code="TAM",
                    theory=True,
                    practical=False,
                    internal=True,
                    theory_max=80,
                    practical_max=0,
                    internal_max=20,
                ),

                Subject(
                    name="English",
                    code="ENG",
                    theory=True,
                    practical=False,
                    internal=True,
                    theory_max=80,
                    practical_max=0,
                    internal_max=20,
                ),

                Subject(
                    name="Mathematics",
                    code="MAT",
                    theory=True,
                    practical=False,
                    internal=True,
                    theory_max=80,
                    practical_max=0,
                    internal_max=20,
                ),

                Subject(
                    name="Physics",
                    code="PHY",
                    theory=True,
                    practical=True,
                    internal=True,
                    theory_max=70,
                    practical_max=20,
                    internal_max=10,
                ),

                Subject(
                    name="Chemistry",
                    code="CHE",
                    theory=True,
                    practical=True,
                    internal=True,
                    theory_max=70,
                    practical_max=20,
                    internal_max=10,
                ),

                Subject(
                    name="Computer Science",
                    code="CS",
                    theory=True,
                    practical=True,
                    internal=True,
                    theory_max=70,
                    practical_max=20,
                    internal_max=10,
                ),

            ])


        # --------------------------------------------------
        # Admin
        # --------------------------------------------------

        admin = (
            db.query(Teacher)
            .filter_by(
                username="admin"
            )
            .first()
        )


        if admin is None:

            db.add(
                Teacher(
                    username="admin",
                    name="Administrator",
                    password_hash=hash_password(
                        "Admin@123"
                    ),
                    role="admin",
                    active=True,
                )
            )

        else:

            # Always make sure admin remains active
            admin.active = True
            admin.role = "admin"


        db.commit()


        # --------------------------------------------------
        # Connect active subjects to all classes
        # --------------------------------------------------

        all_classes = (
            db.query(ClassSection)
            .all()
        )

        all_subjects = (
            db.query(Subject)
            .filter(
                Subject.active == True
            )
            .all()
        )


        for cls in all_classes:

            for subject in all_subjects:

                exists = (
                    db.query(ClassSubject)
                    .filter_by(
                        class_id=cls.id,
                        subject_id=subject.id,
                    )
                    .first()
                )


                if not exists:

                    db.add(
                        ClassSubject(
                            class_id=cls.id,
                            subject_id=subject.id,
                        )
                    )


        db.commit()


    except Exception:

        db.rollback()
        raise

    finally:

        db.close()


seed_database()


# ==========================================================
# CHOICE HELPERS
# ==========================================================


def get_classes():

    db = SessionLocal()

    try:

        rows = (
            db.query(ClassSection)
            .order_by(
                ClassSection.name
            )
            .all()
        )

        return [
            f"{row.id} | {row.name}"
            for row in rows
        ]

    finally:

        db.close()


def get_years():

    db = SessionLocal()

    try:

        rows = (
            db.query(AcademicYear)
            .filter(
                AcademicYear.active == True
            )
            .order_by(
                AcademicYear.name
            )
            .all()
        )

        return [
            f"{row.id} | {row.name}"
            for row in rows
        ]

    finally:

        db.close()


def get_exams():

    db = SessionLocal()

    try:

        rows = (
            db.query(Exam)
            .filter(
                Exam.active == True
            )
            .order_by(
                Exam.id
            )
            .all()
        )

        return [
            f"{row.id} | {row.name}"
            for row in rows
        ]

    finally:

        db.close()


def get_all_exams():

    db = SessionLocal()

    try:

        rows = (
            db.query(Exam)
            .filter(
                Exam.active == True
            )
            .order_by(
                Exam.id
            )
            .all()
        )

        return (
            ["ALL | All Exams"]
            + [
                f"{row.id} | {row.name}"
                for row in rows
            ]
        )

    finally:

        db.close()


def get_subject_choices():

    db = SessionLocal()

    try:

        rows = (
            db.query(Subject)
            .filter(
                Subject.active == True
            )
            .order_by(
                Subject.name
            )
            .all()
        )

        return [
            f"{row.id} | {row.name}"
            for row in rows
        ]

    finally:

        db.close()


def get_subjects_for_class(
    class_value
):

    class_id = parse_id(
        class_value
    )

    if not class_id:

        return gr.Dropdown(
            choices=[],
            value=None,
        )


    db = SessionLocal()

    try:

        subjects = (
            db.query(Subject)
            .join(
                ClassSubject,
                ClassSubject.subject_id
                == Subject.id,
            )
            .filter(
                ClassSubject.class_id
                == class_id,
                Subject.active == True,
            )
            .order_by(
                Subject.name
            )
            .all()
        )

        choices = [
            f"{subject.id} | {subject.name}"
            for subject in subjects
        ]

        return gr.Dropdown(
            choices=choices,
            value=None,
        )

    finally:

        db.close()


def get_student_delete_choices():

    db = SessionLocal()

    try:

        students = (
            db.query(Student)
            .filter(
                Student.active == True
            )
            .order_by(
                Student.roll_no,
                Student.name,
            )
            .all()
        )

        return [
            (
                f"{student.id} | "
                f"{student.roll_no or '-'} | "
                f"{student.name}"
            )
            for student in students
        ]

    finally:

        db.close()


# ==========================================================
# TEACHER CHOICES
# ==========================================================


def get_teacher_delete_choices():

    db = SessionLocal()

    try:

        teachers = (
            db.query(Teacher)
            .filter(
                Teacher.active == True,
                Teacher.username != "admin",
            )
            .order_by(
                Teacher.username
            )
            .all()
        )

        return [
            (
                f"{teacher.id} | "
                f"{teacher.username} | "
                f"{teacher.name}"
            )
            for teacher in teachers
        ]

    finally:

        db.close()


def get_teacher_list():

    db = SessionLocal()

    try:

        teachers = (
            db.query(Teacher)
            .order_by(
                Teacher.username
            )
            .all()
        )

        return [

            [
                teacher.id,
                teacher.username,
                teacher.name,
                teacher.role,
                "Active"
                if teacher.active
                else "Deleted",
            ]

            for teacher in teachers

        ]

    finally:

        db.close()


# ==========================================================
# STUDENT LIST
# ==========================================================


def get_student_list():

    db = SessionLocal()

    try:

        rows = (
            db.query(
                Student,
                ClassSection,
            )
            .join(
                ClassSection,
                Student.class_id
                == ClassSection.id,
            )
            .filter(
                Student.active == True
            )
            .order_by(
                ClassSection.name,
                Student.roll_no,
                Student.name,
            )
            .all()
        )

        return [

            [
                student.id,
                student.admission_no,
                student.roll_no or "",
                student.name,
                cls.name,
            ]

            for student, cls in rows

        ]

    finally:

        db.close()


# ==========================================================
# ACADEMIC YEAR LIST
# ==========================================================


def get_year_list():

    db = SessionLocal()

    try:

        rows = (
            db.query(AcademicYear)
            .order_by(
                AcademicYear.name
            )
            .all()
        )

        return [

            [
                row.id,
                row.name,
                "Active"
                if row.active
                else "Deleted",
            ]

            for row in rows

        ]

    finally:

        db.close()


# ==========================================================
# CLASS LIST
# ==========================================================


def get_class_list():

    db = SessionLocal()

    try:

        rows = (
            db.query(ClassSection)
            .order_by(
                ClassSection.name
            )
            .all()
        )

        return [
            [
                row.id,
                row.name,
            ]
            for row in rows
        ]

    finally:

        db.close()


# ==========================================================
# SUBJECT LIST
# ==========================================================


def get_subject_list():

    db = SessionLocal()

    try:

        rows = (
            db.query(Subject)
            .order_by(
                Subject.name
            )
            .all()
        )

        result = []

        for subject in rows:

            components = []

            if subject.theory:

                components.append(
                    f"Theory ({subject.theory_max})"
                )

            if subject.practical:

                components.append(
                    f"Practical ({subject.practical_max})"
                )

            if subject.internal:

                components.append(
                    f"Internal ({subject.internal_max})"
                )

            result.append([

                subject.id,

                subject.name,

                subject.code,

                " + ".join(
                    components
                ),

                "Active"
                if subject.active
                else "Deleted",

            ])

        return result

    finally:

        db.close()


# ==========================================================
# EXAM LIST
# ==========================================================


def get_exam_list():

    db = SessionLocal()

    try:

        rows = (
            db.query(Exam)
            .order_by(
                Exam.id
            )
            .all()
        )

        return [

            [
                row.id,
                row.name,
                "Active"
                if row.active
                else "Deleted",
            ]

            for row in rows

        ]

    finally:

        db.close()


# ==========================================================
# ADD TEACHER
# ==========================================================


def add_teacher(
    teacher_username,
    teacher_name,
    teacher_password,
):

    username = clean_name(
        teacher_username
    ).lower()

    name = clean_name(
        teacher_name
    )

    password = str(
        teacher_password or ""
    ).strip()


    if not username:

        return (
            "❌ Teacher username is required.",
            get_teacher_list(),
            gr.Dropdown(
                choices=get_teacher_delete_choices(),
                value=None,
            ),
        )


    if not name:

        return (
            "❌ Teacher name is required.",
            get_teacher_list(),
            gr.Dropdown(
                choices=get_teacher_delete_choices(),
                value=None,
            ),
        )


    if not password:

        return (
            "❌ Teacher password is required.",
            get_teacher_list(),
            gr.Dropdown(
                choices=get_teacher_delete_choices(),
                value=None,
            ),
        )


    if len(password) < 6:

        return (
            "❌ Password must contain at least 6 characters.",
            get_teacher_list(),
            gr.Dropdown(
                choices=get_teacher_delete_choices(),
                value=None,
            ),
        )


    db = SessionLocal()

    try:

        existing = (
            db.query(Teacher)
            .filter(
                Teacher.username == username
            )
            .first()
        )


        if existing:

            if not existing.active:

                existing.active = True
                existing.name = name
                existing.password_hash = hash_password(
                    password
                )
                existing.role = "teacher"

                db.commit()

                return (
                    "✅ Teacher restored successfully.",
                    get_teacher_list(),
                    gr.Dropdown(
                        choices=get_teacher_delete_choices(),
                        value=None,
                    ),
                )


            return (
                "❌ This username already exists.",
                get_teacher_list(),
                gr.Dropdown(
                    choices=get_teacher_delete_choices(),
                    value=None,
                ),
            )


        teacher = Teacher(

            username=username,

            name=name,

            password_hash=hash_password(
                password
            ),

            role="teacher",

            active=True,

        )


        db.add(teacher)

        db.commit()


        return (
            "✅ Teacher added successfully.",
            get_teacher_list(),
            gr.Dropdown(
                choices=get_teacher_delete_choices(),
                value=None,
            ),
        )


    except Exception as e:

        db.rollback()

        return (
            "❌ Error: " + str(e),
            get_teacher_list(),
            gr.Dropdown(
                choices=get_teacher_delete_choices(),
                value=None,
            ),
        )

    finally:

        db.close()


# ==========================================================
# DELETE TEACHER
# ==========================================================


def delete_teacher(
    teacher_value
):

    teacher_id = parse_id(
        teacher_value
    )


    if not teacher_id:

        return (
            "❌ Please select a Teacher.",
            get_teacher_list(),
            gr.Dropdown(
                choices=get_teacher_delete_choices(),
                value=None,
            ),
        )


    db = SessionLocal()

    try:

        teacher = db.get(
            Teacher,
            teacher_id,
        )


        if teacher is None:

            return (
                "❌ Teacher not found.",
                get_teacher_list(),
                gr.Dropdown(
                    choices=get_teacher_delete_choices(),
                    value=None,
                ),
            )


        if teacher.username == "admin" or teacher.role == "admin":

            return (
                "❌ Admin account cannot be deleted.",
                get_teacher_list(),
                gr.Dropdown(
                    choices=get_teacher_delete_choices(),
                    value=None,
                ),
            )


        teacher.active = False

        db.commit()


        return (
            "✅ Teacher deleted successfully.",
            get_teacher_list(),
            gr.Dropdown(
                choices=get_teacher_delete_choices(),
                value=None,
            ),
        )


    except Exception as e:

        db.rollback()

        return (
            "❌ Error: " + str(e),
            get_teacher_list(),
            gr.Dropdown(
                choices=get_teacher_delete_choices(),
                value=None,
            ),
        )

    finally:

        db.close()


# ==========================================================
# ADD STUDENT
# ==========================================================


def add_student(
    admission_no,
    roll_no,
    student_name,
    class_value,
):

    admission_no = clean_name(
        admission_no
    )

    roll_no = clean_name(
        roll_no
    )

    student_name = clean_name(
        student_name
    )

    class_id = parse_id(
        class_value
    )


    if not admission_no:

        return (
            "❌ Admission No is required.",
            get_student_list(),
            gr.Dropdown(
                choices=get_student_delete_choices(),
                value=None,
            ),
        )


    if not student_name:

        return (
            "❌ Student Name is required.",
            get_student_list(),
            gr.Dropdown(
                choices=get_student_delete_choices(),
                value=None,
            ),
        )


    if not class_id:

        return (
            "❌ Please select Class.",
            get_student_list(),
            gr.Dropdown(
                choices=get_student_delete_choices(),
                value=None,
            ),
        )


    db = SessionLocal()

    try:

        class_obj = db.get(
            ClassSection,
            class_id,
        )


        if class_obj is None:

            return (
                "❌ Selected Class not found.",
                get_student_list(),
                gr.Dropdown(
                    choices=get_student_delete_choices(),
                    value=None,
                ),
            )


        existing = (
            db.query(Student)
            .filter(
                Student.admission_no
                == admission_no
            )
            .first()
        )


        if existing:

            if not existing.active:

                existing.active = True
                existing.roll_no = roll_no
                existing.name = student_name
                existing.class_id = class_id

                db.commit()

                return (
                    "✅ Student restored successfully.",
                    get_student_list(),
                    gr.Dropdown(
                        choices=get_student_delete_choices(),
                        value=None,
                    ),
                )


            return (
                "❌ This Admission No already exists.",
                get_student_list(),
                gr.Dropdown(
                    choices=get_student_delete_choices(),
                    value=None,
                ),
            )


        student = Student(

            admission_no=admission_no,

            roll_no=roll_no,

            name=student_name,

            class_id=class_id,

            active=True,

        )


        db.add(student)

        db.commit()


        return (
            "✅ Student added successfully.",
            get_student_list(),
            gr.Dropdown(
                choices=get_student_delete_choices(),
                value=None,
            ),
        )


    except Exception as e:

        db.rollback()

        return (
            "❌ Error: " + str(e),
            get_student_list(),
            gr.Dropdown(
                choices=get_student_delete_choices(),
                value=None,
            ),
        )

    finally:

        db.close()


# ==========================================================
# DELETE STUDENT
# ==========================================================


def delete_student(
    student_value
):

    student_id = parse_id(
        student_value
    )


    if not student_id:

        return (
            "❌ Please select a Student.",
            get_student_list(),
            gr.Dropdown(
                choices=get_student_delete_choices(),
                value=None,
            ),
        )


    db = SessionLocal()

    try:

        student = db.get(
            Student,
            student_id,
        )


        if student is None:

            return (
                "❌ Student not found.",
                get_student_list(),
                gr.Dropdown(
                    choices=get_student_delete_choices(),
                    value=None,
                ),
            )


        if not student.active:

            return (
                "❌ Student is already deleted.",
                get_student_list(),
                gr.Dropdown(
                    choices=get_student_delete_choices(),
                    value=None,
                ),
            )


        student.active = False

        db.commit()


        return (
            "✅ Student deleted successfully. "
            "Existing marks are preserved.",
            get_student_list(),
            gr.Dropdown(
                choices=get_student_delete_choices(),
                value=None,
            ),
        )


    except Exception as e:

        db.rollback()

        return (
            "❌ Error: " + str(e),
            get_student_list(),
            gr.Dropdown(
                choices=get_student_delete_choices(),
                value=None,
            ),
        )

    finally:

        db.close()


# ==========================================================
# ADD ACADEMIC YEAR
# ==========================================================


def add_academic_year(
    year_name
):

    year_name = clean_name(
        year_name
    )


    choices = get_years()


    if not year_name:

        return (
            "❌ Academic Year is required.",
            get_year_list(),
            *[
                gr.Dropdown(
                    choices=choices,
                    value=None,
                )
                for _ in range(4)
            ],
        )


    db = SessionLocal()

    try:

        existing = (
            db.query(AcademicYear)
            .filter(
                AcademicYear.name
                == year_name
            )
            .first()
        )


        if existing:

            if not existing.active:

                existing.active = True

                db.commit()

                choices = get_years()

                return (
                    "✅ Academic Year restored.",
                    get_year_list(),
                    *[
                        gr.Dropdown(
                            choices=choices,
                            value=None,
                        )
                        for _ in range(4)
                    ],
                )


            return (
                "❌ Academic Year already exists.",
                get_year_list(),
                *[
                    gr.Dropdown(
                        choices=choices,
                        value=None,
                    )
                    for _ in range(4)
                ],
            )


        db.add(
            AcademicYear(
                name=year_name,
                active=True,
            )
        )

        db.commit()

        choices = get_years()

        return (
            "✅ Academic Year added successfully.",
            get_year_list(),
            *[
                gr.Dropdown(
                    choices=choices,
                    value=None,
                )
                for _ in range(4)
            ],
        )


    except Exception as e:

        db.rollback()

        choices = get_years()

        return (
            "❌ Error: " + str(e),
            get_year_list(),
            *[
                gr.Dropdown(
                    choices=choices,
                    value=None,
                )
                for _ in range(4)
            ],
        )

    finally:

        db.close()


# ==========================================================
# DELETE ACADEMIC YEAR
# ==========================================================


def delete_academic_year(
    year_value
):

    year_id = parse_id(
        year_value
    )

    choices = get_years()


    if not year_id:

        return (
            "❌ Please select an Academic Year.",
            get_year_list(),
            *[
                gr.Dropdown(
                    choices=choices,
                    value=None,
                )
                for _ in range(4)
            ],
        )


    db = SessionLocal()

    try:

        year = db.get(
            AcademicYear,
            year_id,
        )


        if year is None:

            return (
                "❌ Academic Year not found.",
                get_year_list(),
                *[
                    gr.Dropdown(
                        choices=choices,
                        value=None,
                    )
                    for _ in range(4)
                ],
            )


        year.active = False

        db.commit()

        choices = get_years()

        return (
            "✅ Academic Year deleted. "
            "Existing marks are preserved.",
            get_year_list(),
            *[
                gr.Dropdown(
                    choices=choices,
                    value=None,
                )
                for _ in range(4)
            ],
        )


    except Exception as e:

        db.rollback()

        choices = get_years()

        return (
            "❌ Error: " + str(e),
            get_year_list(),
            *[
                gr.Dropdown(
                    choices=choices,
                    value=None,
                )
                for _ in range(4)
            ],
        )

    finally:

        db.close()


# ==========================================================
# ADD CLASS
# ==========================================================


def add_class(
    class_name
):

    class_name = clean_name(
        class_name
    )

    choices = get_classes()


    if not class_name:

        return (
            "❌ Class name is required.",
            get_class_list(),
            *[
                gr.Dropdown(
                    choices=choices,
                    value=None,
                )
                for _ in range(5)
            ],
        )


    db = SessionLocal()

    try:

        existing = (
            db.query(ClassSection)
            .filter(
                ClassSection.name
                == class_name
            )
            .first()
        )


        if existing:

            choices = get_classes()

            return (
                "❌ Class already exists.",
                get_class_list(),
                *[
                    gr.Dropdown(
                        choices=choices,
                        value=None,
                    )
                    for _ in range(5)
                ],
            )


        new_class = ClassSection(
            name=class_name
        )

        db.add(new_class)

        db.flush()


        subjects = (
            db.query(Subject)
            .filter(
                Subject.active == True
            )
            .all()
        )


        for subject in subjects:

            exists = (
                db.query(ClassSubject)
                .filter_by(
                    class_id=new_class.id,
                    subject_id=subject.id,
                )
                .first()
            )


            if not exists:

                db.add(
                    ClassSubject(
                        class_id=new_class.id,
                        subject_id=subject.id,
                    )
                )


        db.commit()

        choices = get_classes()

        return (
            "✅ Class added successfully.",
            get_class_list(),
            *[
                gr.Dropdown(
                    choices=choices,
                    value=None,
                )
                for _ in range(5)
            ],
        )


    except Exception as e:

        db.rollback()

        choices = get_classes()

        return (
            "❌ Error: " + str(e),
            get_class_list(),
            *[
                gr.Dropdown(
                    choices=choices,
                    value=None,
                )
                for _ in range(5)
            ],
        )

    finally:

        db.close()


# ==========================================================
# DELETE CLASS
# ==========================================================


def delete_class(
    class_value
):

    class_id = parse_id(
        class_value
    )

    choices = get_classes()


    if not class_id:

        return (
            "❌ Please select a Class.",
            get_class_list(),
            *[
                gr.Dropdown(
                    choices=choices,
                    value=None,
                )
                for _ in range(5)
            ],
        )


    db = SessionLocal()

    try:

        cls = db.get(
            ClassSection,
            class_id,
        )


        if cls is None:

            return (
                "❌ Class not found.",
                get_class_list(),
                *[
                    gr.Dropdown(
                        choices=choices,
                        value=None,
                    )
                    for _ in range(5)
                ],
            )


        active_students = (
            db.query(Student)
            .filter(
                Student.class_id == class_id,
                Student.active == True,
            )
            .count()
        )


        if active_students > 0:

            return (
                "❌ Cannot delete this Class because "
                "it still has active students.",
                get_class_list(),
                *[
                    gr.Dropdown(
                        choices=choices,
                        value=None,
                    )
                    for _ in range(5)
                ],
            )


        db.query(
            ClassSubject
        ).filter(
            ClassSubject.class_id
            == class_id
        ).delete(
            synchronize_session=False
        )


        db.delete(cls)

        db.commit()

        choices = get_classes()

        return (
            "✅ Class deleted successfully.",
            get_class_list(),
            *[
                gr.Dropdown(
                    choices=choices,
                    value=None,
                )
                for _ in range(5)
            ],
        )


    except Exception as e:

        db.rollback()

        choices = get_classes()

        return (
            "❌ Error: " + str(e),
            get_class_list(),
            *[
                gr.Dropdown(
                    choices=choices,
                    value=None,
                )
                for _ in range(5)
            ],
        )

    finally:

        db.close()


# ==========================================================
# ADD SUBJECT
# ==========================================================


def add_subject(
    subject_name,
    subject_code,
    theory,
    practical,
    internal,
    theory_max,
    practical_max,
    internal_max,
):

    subject_name = clean_name(
        subject_name
    )

    subject_code = clean_name(
        subject_code
    ).upper()


    if not subject_name:

        return (
            "❌ Subject name is required.",
            get_subject_list(),
            gr.Dropdown(
                choices=get_subject_choices(),
                value=None,
            ),
        )


    if not subject_code:

        return (
            "❌ Subject code is required.",
            get_subject_list(),
            gr.Dropdown(
                choices=get_subject_choices(),
                value=None,
            ),
        )


    if not any([
        theory,
        practical,
        internal,
    ]):

        return (
            "❌ Select at least one component.",
            get_subject_list(),
            gr.Dropdown(
                choices=get_subject_choices(),
                value=None,
            ),
        )


    theory_max = safe_int(
        theory_max
    )

    practical_max = safe_int(
        practical_max
    )

    internal_max = safe_int(
        internal_max
    )


    if theory and theory_max <= 0:

        return (
            "❌ Theory maximum must be greater than 0.",
            get_subject_list(),
            gr.Dropdown(
                choices=get_subject_choices(),
                value=None,
            ),
        )


    if practical and practical_max <= 0:

        return (
            "❌ Practical maximum must be greater than 0.",
            get_subject_list(),
            gr.Dropdown(
                choices=get_subject_choices(),
                value=None,
            ),
        )


    if internal and internal_max <= 0:

        return (
            "❌ Internal maximum must be greater than 0.",
            get_subject_list(),
            gr.Dropdown(
                choices=get_subject_choices(),
                value=None,
            ),
        )


    db = SessionLocal()

    try:

        existing = (
            db.query(Subject)
            .filter(
                Subject.code
                == subject_code
            )
            .first()
        )


        if existing:

            if not existing.active:

                existing.active = True
                existing.name = subject_name
                existing.theory = bool(theory)
                existing.practical = bool(practical)
                existing.internal = bool(internal)
                existing.theory_max = theory_max
                existing.practical_max = practical_max
                existing.internal_max = internal_max

                subject_id = existing.id

            else:

                return (
                    "❌ Subject code already exists.",
                    get_subject_list(),
                    gr.Dropdown(
                        choices=get_subject_choices(),
                        value=None,
                    ),
                )

        else:

            subject = Subject(

                name=subject_name,

                code=subject_code,

                theory=bool(theory),

                practical=bool(practical),

                internal=bool(internal),

                theory_max=theory_max,

                practical_max=practical_max,

                internal_max=internal_max,

                active=True,

            )

            db.add(subject)

            db.flush()

            subject_id = subject.id


        classes = (
            db.query(ClassSection)
            .all()
        )


        for cls in classes:

            exists = (
                db.query(ClassSubject)
                .filter_by(
                    class_id=cls.id,
                    subject_id=subject_id,
                )
                .first()
            )


            if not exists:

                db.add(
                    ClassSubject(
                        class_id=cls.id,
                        subject_id=subject_id,
                    )
                )


        db.commit()


        return (
            "✅ Subject added successfully.",
            get_subject_list(),
            gr.Dropdown(
                choices=get_subject_choices(),
                value=None,
            ),
        )


    except Exception as e:

        db.rollback()

        return (
            "❌ Error: " + str(e),
            get_subject_list(),
            gr.Dropdown(
                choices=get_subject_choices(),
                value=None,
            ),
        )

    finally:

        db.close()


# ==========================================================
# DELETE SUBJECT
# ==========================================================


def delete_subject(
    subject_value
):

    subject_id = parse_id(
        subject_value
    )


    if not subject_id:

        return (
            "❌ Please select a Subject.",
            get_subject_list(),
            gr.Dropdown(
                choices=get_subject_choices(),
                value=None,
            ),
        )


    db = SessionLocal()

    try:

        subject = db.get(
            Subject,
            subject_id,
        )


        if subject is None:

            return (
                "❌ Subject not found.",
                get_subject_list(),
                gr.Dropdown(
                    choices=get_subject_choices(),
                    value=None,
                ),
            )


        subject.active = False

        db.commit()


        return (
            "✅ Subject deleted. Existing marks are preserved.",
            get_subject_list(),
            gr.Dropdown(
                choices=get_subject_choices(),
                value=None,
            ),
        )


    except Exception as e:

        db.rollback()

        return (
            "❌ Error: " + str(e),
            get_subject_list(),
            gr.Dropdown(
                choices=get_subject_choices(),
                value=None,
            ),
        )

    finally:

        db.close()


# ==========================================================
# ADD EXAM
# ==========================================================


def add_exam(
    exam_name
):

    exam_name = clean_name(
        exam_name
    )


    if not exam_name:

        choices = get_exams()

        return (
            "❌ Exam name is required.",
            get_exam_list(),
            *[
                gr.Dropdown(
                    choices=choices,
                    value=None,
                )
                for _ in range(4)
            ],
        )


    db = SessionLocal()

    try:

        existing = (
            db.query(Exam)
            .filter(
                Exam.name
                == exam_name
            )
            .first()
        )


        if existing:

            if not existing.active:

                existing.active = True

                db.commit()

                choices = get_exams()

                return (
                    "✅ Exam restored successfully.",
                    get_exam_list(),
                    *[
                        gr.Dropdown(
                            choices=choices,
                            value=None,
                        )
                        for _ in range(4)
                    ],
                )


            choices = get_exams()

            return (
                "❌ Exam already exists.",
                get_exam_list(),
                *[
                    gr.Dropdown(
                        choices=choices,
                        value=None,
                    )
                    for _ in range(4)
                ],
            )


        db.add(
            Exam(
                name=exam_name,
                active=True,
            )
        )

        db.commit()

        choices = get_exams()

        return (
            "✅ Exam added successfully.",
            get_exam_list(),
            *[
                gr.Dropdown(
                    choices=choices,
                    value=None,
                )
                for _ in range(4)
            ],
        )


    except Exception as e:

        db.rollback()

        choices = get_exams()

        return (
            "❌ Error: " + str(e),
            get_exam_list(),
            *[
                gr.Dropdown(
                    choices=choices,
                    value=None,
                )
                for _ in range(4)
            ],
        )

    finally:

        db.close()


# ==========================================================
# DELETE EXAM
# ==========================================================


def delete_exam(
    exam_value
):

    exam_id = parse_id(
        exam_value
    )


    if not exam_id:

        choices = get_exams()

        return (
            "❌ Please select an Exam.",
            get_exam_list(),
            *[
                gr.Dropdown(
                    choices=choices,
                    value=None,
                )
                for _ in range(4)
            ],
        )


    db = SessionLocal()

    try:

        exam = db.get(
            Exam,
            exam_id,
        )


        if exam is None:

            choices = get_exams()

            return (
                "❌ Exam not found.",
                get_exam_list(),
                *[
                    gr.Dropdown(
                        choices=choices,
                        value=None,
                    )
                    for _ in range(4)
                ],
            )


        exam.active = False

        db.commit()

        choices = get_exams()

        return (
            "✅ Exam deleted. Existing marks are preserved.",
            get_exam_list(),
            *[
                gr.Dropdown(
                    choices=choices,
                    value=None,
                )
                for _ in range(4)
            ],
        )


    except Exception as e:

        db.rollback()

        choices = get_exams()

        return (
            "❌ Error: " + str(e),
            get_exam_list(),
            *[
                gr.Dropdown(
                    choices=choices,
                    value=None,
                )
                for _ in range(4)
            ],
        )

    finally:

        db.close()


# ==========================================================
# MARK ENTRY - LOAD
# ==========================================================


def load_marks(
    class_value,
    year_value,
    exam_value,
    subject_value,
):

    class_id = parse_id(
        class_value
    )

    year_id = parse_id(
        year_value
    )

    exam_id = parse_id(
        exam_value
    )

    subject_id = parse_id(
        subject_value
    )


    if not class_id:

        return (
            [],
            "❌ Please select Class.",
        )


    if not year_id:

        return (
            [],
            "❌ Please select Academic Year.",
        )


    if not exam_id:

        return (
            [],
            "❌ Please select Exam.",
        )


    if not subject_id:

        return (
            [],
            "❌ Please select Subject.",
        )


    db = SessionLocal()

    try:

        subject = db.get(
            Subject,
            subject_id,
        )


        if subject is None or not subject.active:

            return (
                [],
                "❌ Subject not found.",
            )


        students = (
            db.query(Student)
            .filter(
                Student.class_id == class_id,
                Student.active == True,
            )
            .order_by(
                Student.roll_no,
                Student.name,
            )
            .all()
        )


        headers = [
            "ID",
            "Roll No",
            "Student Name",
        ]


        if subject.theory:

            headers.append(
                f"Theory / {subject.theory_max}"
            )


        if subject.practical:

            headers.append(
                f"Practical / {subject.practical_max}"
            )


        if subject.internal:

            headers.append(
                f"Internal / {subject.internal_max}"
            )


        headers.append(
            "Total"
        )


        rows = []


        for student in students:

            mark = (
                db.query(Mark)
                .filter_by(
                    academic_year_id=year_id,
                    exam_id=exam_id,
                    student_id=student.id,
                    subject_id=subject_id,
                )
                .first()
            )


            theory = (
                mark.theory
                if mark
                else 0
            )

            practical = (
                mark.practical
                if mark
                else 0
            )

            internal = (
                mark.internal
                if mark
                else 0
            )


            total = (
                theory
                + practical
                + internal
            )


            row = [

                student.id,

                student.roll_no or "",

                student.name,

            ]


            if subject.theory:

                row.append(
                    theory
                )


            if subject.practical:

                row.append(
                    practical
                )


            if subject.internal:

                row.append(
                    internal
                )


            row.append(
                total
            )


            rows.append(row)


        pattern = []


        if subject.theory:

            pattern.append(
                f"Theory: {subject.theory_max}"
            )


        if subject.practical:

            pattern.append(
                f"Practical: {subject.practical_max}"
            )


        if subject.internal:

            pattern.append(
                f"Internal: {subject.internal_max}"
            )


        dataframe = pd.DataFrame(
            rows,
            columns=headers,
        )


        return (
            dataframe,
            "**Mark Pattern:** "
            + " + ".join(pattern),
        )


    except Exception as e:

        return (
            [],
            "❌ Error while loading marks: "
            + str(e),
        )

    finally:

        db.close()


# ==========================================================
# MARK ENTRY - SAVE
# ==========================================================


def save_marks(
    class_value,
    year_value,
    exam_value,
    subject_value,
    table_data,
):

    class_id = parse_id(
        class_value
    )

    year_id = parse_id(
        year_value
    )

    exam_id = parse_id(
        exam_value
    )

    subject_id = parse_id(
        subject_value
    )


    if not class_id:

        return "❌ Please select Class."


    if not year_id:

        return "❌ Please select Academic Year."


    if not exam_id:

        return "❌ Please select Exam."


    if not subject_id:

        return "❌ Please select Subject."


    if table_data is None:

        return (
            "❌ No student data found. "
            "Please click Load Students."
        )


    db = SessionLocal()

    try:

        subject = (
            db.query(Subject)
            .filter(
                Subject.id == subject_id,
                Subject.active == True,
            )
            .first()
        )


        if subject is None:

            return "❌ Subject not found."


        if isinstance(
            table_data,
            pd.DataFrame
        ):

            rows = table_data.values.tolist()

        else:

            rows = table_data


        if not rows:

            return (
                "❌ No student rows found."
            )


        saved_count = 0


        for row in rows:

            if row is None:

                continue


            if len(row) < 3:

                continue


            try:

                student_id = int(
                    float(row[0])
                )

            except Exception:

                continue


            student = (
                db.query(Student)
                .filter(
                    Student.id == student_id,
                    Student.class_id == class_id,
                    Student.active == True,
                )
                .first()
            )


            if student is None:

                continue


            position = 3


            theory = 0
            practical = 0
            internal = 0


            if subject.theory:

                if len(row) > position:

                    theory = safe_int(
                        row[position]
                    )

                position += 1


            if subject.practical:

                if len(row) > position:

                    practical = safe_int(
                        row[position]
                    )

                position += 1


            if subject.internal:

                if len(row) > position:

                    internal = safe_int(
                        row[position]
                    )

                position += 1


            if theory < 0 or theory > subject.theory_max:

                return (
                    f"❌ Invalid Theory mark for "
                    f"{student.name}. "
                    f"Maximum is {subject.theory_max}."
                )


            if (
                subject.practical
                and (
                    practical < 0
                    or practical > subject.practical_max
                )
            ):

                return (
                    f"❌ Invalid Practical mark for "
                    f"{student.name}. "
                    f"Maximum is {subject.practical_max}."
                )


            if (
                subject.internal
                and (
                    internal < 0
                    or internal > subject.internal_max
                )
            ):

                return (
                    f"❌ Invalid Internal mark for "
                    f"{student.name}. "
                    f"Maximum is {subject.internal_max}."
                )


            total = (
                theory
                + practical
                + internal
            )


            mark = (
                db.query(Mark)
                .filter_by(
                    academic_year_id=year_id,
                    exam_id=exam_id,
                    student_id=student_id,
                    subject_id=subject_id,
                )
                .first()
            )


            if mark is None:

                mark = Mark(

                    academic_year_id=year_id,

                    exam_id=exam_id,

                    student_id=student_id,

                    subject_id=subject_id,

                )

                db.add(mark)


            mark.theory = theory

            mark.practical = practical

            mark.internal = internal

            mark.total = total

            mark.updated_at = datetime.utcnow()


            saved_count += 1


        if saved_count == 0:

            return (
                "❌ No valid student rows found."
            )


        db.commit()


        return (
            f"✅ Marks saved successfully for "
            f"{saved_count} student(s)."
        )


    except Exception as e:

        db.rollback()

        return (
            "❌ Error while saving marks: "
            + str(e)
        )

    finally:

        db.close()



# ==========================================================
# MOBILE MARK ENTRY
# ==========================================================


def save_mobile_marks(
    class_value,
    year_value,
    exam_value,
    subject_value,
    *mark_values,
):
    """Save marks entered through mobile-friendly Number inputs."""

    class_id = parse_id(class_value)
    year_id = parse_id(year_value)
    exam_id = parse_id(exam_value)
    subject_id = parse_id(subject_value)

    if not class_id:
        return "❌ Please select Class."
    if not year_id:
        return "❌ Please select Academic Year."
    if not exam_id:
        return "❌ Please select Exam."
    if not subject_id:
        return "❌ Please select Subject."

    db = SessionLocal()

    try:
        subject = (
            db.query(Subject)
            .filter(
                Subject.id == subject_id,
                Subject.active == True,
            )
            .first()
        )

        if subject is None:
            return "❌ Subject not found."

        students = (
            db.query(Student)
            .filter(
                Student.class_id == class_id,
                Student.active == True,
            )
            .order_by(
                Student.roll_no,
                Student.name,
            )
            .all()
        )

        components = []

        if subject.theory:
            components.append("theory")
        if subject.practical:
            components.append("practical")
        if subject.internal:
            components.append("internal")

        expected = len(students) * len(components)

        if len(mark_values) != expected:
            return (
                "❌ Student list changed. "
                "Please select the Class/Subject again."
            )

        rows = []
        position = 0

        for student in students:
            row = [
                student.id,
                student.roll_no or "",
                student.name,
            ]

            for _ in components:
                row.append(safe_int(mark_values[position]))
                position += 1

            row.append(0)
            rows.append(row)

        # Reuse the existing, fully validated database save routine.
        return save_marks(
            class_value,
            year_value,
            exam_value,
            subject_value,
            rows,
        )

    except Exception as e:
        return "❌ Error while saving marks: " + str(e)

    finally:
        db.close()

# ==========================================================
# CONSOLIDATED REPORT
# ONE STUDENT = ONE ROW
# CLASS-WISE ACTUAL MARK-ENTERED SUBJECTS ONLY
# EXAM-WISE HORIZONTAL COLUMNS
# ZERO / UNUSED EXAMS AND SUBJECTS ARE HIDDEN
# ==========================================================


def generate_consolidated_report(
    year_value,
    class_value,
    exam_value,
):

    year_id = parse_id(year_value)
    class_id = parse_id(class_value)

    if not year_id:
        return (
            "❌ Please select Academic Year.",
            pd.DataFrame(),
        )

    if not class_id:
        return (
            "❌ Please select Class.",
            pd.DataFrame(),
        )

    if not exam_value:
        return (
            "❌ Please select Exam.",
            pd.DataFrame(),
        )

    db = SessionLocal()

    try:

        # --------------------------------------------------
        # CLASS
        # --------------------------------------------------

        class_obj = (
            db.query(ClassSection)
            .filter(
                ClassSection.id == class_id
            )
            .first()
        )

        if class_obj is None:
            return (
                "❌ Class not found.",
                pd.DataFrame(),
            )


        # --------------------------------------------------
        # ACTIVE STUDENTS
        # --------------------------------------------------

        students = (
            db.query(Student)
            .filter(
                Student.class_id == class_id,
                Student.active == True,
            )
            .order_by(
                Student.roll_no,
                Student.name,
            )
            .all()
        )

        if not students:
            return (
                "❌ No students found in this class.",
                pd.DataFrame(),
            )


        # --------------------------------------------------
        # ALL SUBJECTS ASSIGNED TO THIS CLASS
        #
        # IMPORTANT:
        # These are only the possible subjects.
        # We will later keep ONLY subjects which
        # actually have marks entered for this class.
        # --------------------------------------------------

        class_subjects = (
            db.query(Subject)
            .join(
                ClassSubject,
                ClassSubject.subject_id
                == Subject.id,
            )
            .filter(
                ClassSubject.class_id == class_id,
                Subject.active == True,
            )
            .order_by(
                Subject.id
            )
            .all()
        )

        if not class_subjects:
            return (
                "❌ No subjects found for this class.",
                pd.DataFrame(),
            )


        # --------------------------------------------------
        # FIND SELECTED EXAMS
        # --------------------------------------------------

        if str(exam_value).startswith("ALL"):

            all_exams = (
                db.query(Exam)
                .filter(
                    Exam.active == True
                )
                .order_by(
                    Exam.id
                )
                .all()
            )

        else:

            exam_id = parse_id(exam_value)

            exam_obj = (
                db.query(Exam)
                .filter(
                    Exam.id == exam_id,
                    Exam.active == True,
                )
                .first()
            )

            if exam_obj is None:
                return (
                    "❌ Exam not found.",
                    pd.DataFrame(),
                )

            all_exams = [exam_obj]


        # ==================================================
        # STEP 1
        # FIND EXAMS WHICH ACTUALLY HAVE MARKS
        # FOR THIS CLASS
        # ==================================================

        exams = []

        for exam in all_exams:

            exam_has_marks = False

            for student in students:

                for subject in class_subjects:

                    mark = (
                        db.query(Mark)
                        .filter_by(
                            academic_year_id=year_id,
                            exam_id=exam.id,
                            student_id=student.id,
                            subject_id=subject.id,
                        )
                        .first()
                    )

                    if mark is None:
                        continue


                    theory = safe_int(
                        mark.theory
                    )

                    practical = safe_int(
                        mark.practical
                    )

                    internal = safe_int(
                        mark.internal
                    )


                    # --------------------------------------
                    # ONLY NON-ZERO MARK COUNTS
                    # --------------------------------------

                    if (
                        theory != 0
                        or practical != 0
                        or internal != 0
                    ):

                        exam_has_marks = True
                        break


                if exam_has_marks:
                    break


            # ------------------------------------------
            # EXAM WILL BE SHOWN ONLY IF
            # AT LEAST ONE MARK EXISTS
            # ------------------------------------------

            if exam_has_marks:
                exams.append(exam)


        # --------------------------------------------------
        # NO EXAM HAS MARKS
        # --------------------------------------------------

        if not exams:

            return (
                "❌ No marks have been entered for this "
                "Class / Academic Year.",
                pd.DataFrame(),
            )


        # ==================================================
        # STEP 2
        # FIND ACTUAL SUBJECTS FOR EACH EXAM
        #
        # VERY IMPORTANT:
        #
        # A1 may have:
        # Tamil
        # English
        # Computer Science
        #
        # B1 may have:
        # Tamil
        # English
        # Computer Applications
        #
        # We DO NOT hard-code these names.
        #
        # The Subject Name comes directly from database.
        #
        # Only subjects which actually contain
        # NON-ZERO marks for this class + exam
        # are included.
        # ==================================================

        exam_subjects = {}

        for exam in exams:

            active_subjects = []

            for subject in class_subjects:

                subject_has_marks = False

                for student in students:

                    mark = (
                        db.query(Mark)
                        .filter_by(
                            academic_year_id=year_id,
                            exam_id=exam.id,
                            student_id=student.id,
                            subject_id=subject.id,
                        )
                        .first()
                    )

                    if mark is None:
                        continue


                    theory = safe_int(
                        mark.theory
                    )

                    practical = safe_int(
                        mark.practical
                    )

                    internal = safe_int(
                        mark.internal
                    )


                    if (
                        theory != 0
                        or practical != 0
                        or internal != 0
                    ):

                        subject_has_marks = True
                        break


                if subject_has_marks:

                    active_subjects.append(
                        subject
                    )


            # ------------------------------------------
            # STORE ONLY ACTUAL MARK-ENTERED SUBJECTS
            # ------------------------------------------

            if active_subjects:

                exam_subjects[
                    exam.id
                ] = active_subjects


        # --------------------------------------------------
        # REMOVE EXAMS WHICH HAVE NO ACTIVE SUBJECTS
        # --------------------------------------------------

        exams = [
            exam
            for exam in exams
            if exam.id in exam_subjects
            and exam_subjects[exam.id]
        ]


        if not exams:

            return (
                "❌ No valid marks found for this "
                "Class / Academic Year.",
                pd.DataFrame(),
            )


        # ==================================================
        # STEP 3
        # BUILD ONE ROW PER STUDENT
        # ==================================================

        report_rows = []


        for student in students:

            row = {

                "Roll No":
                    student.roll_no or "",

                "Student Name":
                    student.name,

            }


            # --------------------------------------------------
            # EACH EXAM GOES HORIZONTALLY
            # --------------------------------------------------

            for exam in exams:

                exam_total = 0


                # ------------------------------------------
                # ONLY ACTUAL SUBJECTS FOR THIS EXAM
                # ------------------------------------------

                active_subjects = exam_subjects.get(
                    exam.id,
                    []
                )


                for subject in active_subjects:

                    mark = (
                        db.query(Mark)
                        .filter_by(
                            academic_year_id=year_id,
                            exam_id=exam.id,
                            student_id=student.id,
                            subject_id=subject.id,
                        )
                        .first()
                    )


                    theory = (
                        safe_int(mark.theory)
                        if mark
                        else 0
                    )

                    practical = (
                        safe_int(mark.practical)
                        if mark
                        else 0
                    )

                    internal = (
                        safe_int(mark.internal)
                        if mark
                        else 0
                    )


                    subject_total = (
                        theory
                        + practical
                        + internal
                    )


                    exam_total += subject_total


                    # --------------------------------------------------
                    # SUBJECT NAME
                    #
                    # DATABASE SUBJECT NAME IS USED DIRECTLY
                    # --------------------------------------------------

                    subject_name = str(
                        subject.name
                    ).strip()


                    # --------------------------------------------------
                    # SUBJECT COMPONENTS
                    # --------------------------------------------------

                    if subject.theory:

                        row[
                            f"{exam.name} | "
                            f"{subject_name} | THE"
                        ] = theory


                    if subject.practical:

                        row[
                            f"{exam.name} | "
                            f"{subject_name} | PRA"
                        ] = practical


                    if subject.internal:

                        row[
                            f"{exam.name} | "
                            f"{subject_name} | INT"
                        ] = internal


                    # --------------------------------------------------
                    # SUBJECT SUBTOTAL
                    # --------------------------------------------------

                    row[
                        f"{exam.name} | "
                        f"{subject_name} | TOT"
                    ] = subject_total


                # --------------------------------------------------
                # EXAM TOTAL
                #
                # Example:
                # QUARTER TOTAL
                # HALF YEARLY TOTAL
                # ANNUAL TOTAL
                #
                # This is the total ONLY FOR THAT EXAM.
                # No final overall / annual total is added.
                # --------------------------------------------------

                row[
                    f"{exam.name} | TOTAL"
                ] = exam_total


            # --------------------------------------------------
            # ONE STUDENT = ONE ROW
            # --------------------------------------------------

            report_rows.append(
                row
            )


        # ==================================================
        # STEP 4
        # CREATE COLUMN ORDER
        # ==================================================

        columns = [
            "Roll No",
            "Student Name",
        ]


        for exam in exams:

            active_subjects = exam_subjects.get(
                exam.id,
                []
            )


            for subject in active_subjects:

                subject_name = str(
                    subject.name
                ).strip()


                if subject.theory:

                    columns.append(
                        f"{exam.name} | "
                        f"{subject_name} | THE"
                    )


                if subject.practical:

                    columns.append(
                        f"{exam.name} | "
                        f"{subject_name} | PRA"
                    )


                if subject.internal:

                    columns.append(
                        f"{exam.name} | "
                        f"{subject_name} | INT"
                    )


                # ------------------------------------------
                # SUBJECT TOTAL
                # ------------------------------------------

                columns.append(
                    f"{exam.name} | "
                    f"{subject_name} | TOT"
                )


            # ------------------------------------------
            # EXAM TOTAL
            # ------------------------------------------

            columns.append(
                f"{exam.name} | TOTAL"
            )


        # ==================================================
        # STEP 5
        # DATAFRAME
        # ==================================================

        report_df = pd.DataFrame(
            report_rows,
            columns=columns,
        )


        # ==================================================
        # ACADEMIC YEAR NAME
        # ==================================================

        year_text = str(
            year_value
        )

        if "|" in year_text:

            year_name = (
                year_text
                .split("|", 1)[1]
                .strip()
            )

        else:

            year_name = year_text


        # ==================================================
        # EXAM NAMES
        # ==================================================

        exam_names = ", ".join(
            str(exam.name)
            for exam in exams
        )


        # ==================================================
        # SUBJECT NAMES
        # ==================================================

        all_subject_names = []

        for exam in exams:

            for subject in exam_subjects.get(
                exam.id,
                []
            ):

                subject_name = str(
                    subject.name
                ).strip()

                if subject_name not in all_subject_names:

                    all_subject_names.append(
                        subject_name
                    )


        subject_names_text = ", ".join(
            all_subject_names
        )


        # ==================================================
        # REPORT MESSAGE
        # ==================================================

        message = (

            "### 📊 Consolidated Mark List\n\n"

            f"**Academic Year:** {year_name}\n\n"

            f"**Class:** {class_obj.name}\n\n"

            f"**Students:** {len(students)}\n\n"

            f"**Subjects with Marks:** "
            f"{subject_names_text}\n\n"

            f"**Exams with Marks:** "
            f"{exam_names}\n\n"

            "### 📌 Structure\n\n"

            "**One Student = One Row**\n\n"

            "**Exam → Subject → "
            "THE / PRA / INT / TOT → "
            "Exam TOTAL**\n\n"

            "⚠️ Only subjects with entered "
            "non-zero marks are shown.\n\n"

            "⚠️ Exams with no entered marks "
            "are automatically hidden.\n\n"

            "⚠️ Subject names are taken directly "
            "from the database for the selected class."

        )


        return (
            message,
            report_df,
        )


    except Exception as e:

        return (
            "❌ Error while generating report: "
            + str(e),
            pd.DataFrame(),
        )

    finally:

        db.close()

# ==========================================================
# EXCEL DOWNLOAD HELPERS
# ==========================================================


def dataframe_to_excel(
    dataframe,
    filename_prefix,
):

    if dataframe is None:
        return None

    try:

        if isinstance(dataframe, pd.DataFrame):
            df = dataframe.copy()
        else:
            df = pd.DataFrame(dataframe)

    except Exception as e:

        print("DataFrame conversion error:", e)
        return None

    if df.empty:
        print("Excel error: DataFrame is empty")
        return None

    safe_prefix = clean_name(
        filename_prefix
    )

    if not safe_prefix:
        safe_prefix = "school_marks"

    temp_path = os.path.join(
        tempfile.gettempdir(),
        safe_prefix
        + "_"
        + datetime.now().strftime(
            "%Y%m%d_%H%M%S_%f"
        )
        + ".xlsx"
    )

    try:

        with pd.ExcelWriter(
            temp_path,
            engine="openpyxl",
        ) as writer:

            df.to_excel(
                writer,
                index=False,
                sheet_name="Marks",
            )

        print(
            "Excel file created:",
            temp_path
        )

        if os.path.exists(temp_path):

            return temp_path

        return None

    except Exception as e:

        print(
            "Excel creation error:",
            e
        )

        try:

            if os.path.exists(temp_path):
                os.remove(temp_path)

        except Exception:
            pass

        return None

# ==========================================================
# VIEW MARKS EXCEL
# ==========================================================


def download_view_marks_excel(
    table_data
):

    return dataframe_to_excel(
        table_data,
        "view_marks",
    )


# ==========================================================
# CONSOLIDATED EXCEL
# ==========================================================


# ==========================================================
# CONSOLIDATED EXCEL
# ONE STUDENT = ONE ROW
# EXAM -> SUBJECT -> COMPONENTS
# QUARTER -> HALF -> ANNUAL
# SUBJECT SUBTOTAL = TOT
# QUARTER / HALF EXAM TOTAL
# ANNUAL EXAM TOTAL HIDDEN
# ==========================================================


def download_consolidated_excel(
    table_data
):

    if table_data is None:
        return None


    # ------------------------------------------------------
    # CONVERT TO DATAFRAME
    # ------------------------------------------------------

    try:

        if isinstance(
            table_data,
            pd.DataFrame
        ):

            dataframe = table_data.copy()

        else:

            dataframe = pd.DataFrame(
                table_data
            )

    except Exception as e:

        print(
            "DataFrame conversion error:",
            e
        )

        return None


    if dataframe.empty:
        return None


    # ------------------------------------------------------
    # TEMP FILE
    # ------------------------------------------------------

    temp_path = os.path.join(

        tempfile.gettempdir(),

        "consolidated_mark_list_"
        + datetime.now().strftime(
            "%Y%m%d_%H%M%S_%f"
        )
        + ".xlsx"

    )


    try:

        from openpyxl import Workbook

        from openpyxl.styles import (
            Font,
            Alignment,
            Border,
            Side,
        )

        from openpyxl.utils import (
            get_column_letter
        )


        # --------------------------------------------------
        # WORKBOOK
        # --------------------------------------------------

        wb = Workbook()

        ws = wb.active

        ws.title = "Consolidated Marks"


        # --------------------------------------------------
        # BORDER
        # --------------------------------------------------

        thin_border = Border(

            left=Side(
                style="thin"
            ),

            right=Side(
                style="thin"
            ),

            top=Side(
                style="thin"
            ),

            bottom=Side(
                style="thin"
            ),

        )


        # --------------------------------------------------
        # DATA COLUMNS
        # --------------------------------------------------

        dataframe_columns = list(
            dataframe.columns
        )


        if len(dataframe_columns) < 2:

            return None


        # --------------------------------------------------
        # FIXED COLUMNS
        # --------------------------------------------------

        ws.merge_cells(

            start_row=1,

            start_column=1,

            end_row=3,

            end_column=1,

        )

        ws.cell(
            row=1,
            column=1,
            value="Roll No",
        )


        ws.merge_cells(

            start_row=1,

            start_column=2,

            end_row=3,

            end_column=2,

        )

        ws.cell(
            row=1,
            column=2,
            value="Student Name",
        )


        # --------------------------------------------------
        # EXAM COLUMNS
        # --------------------------------------------------

        mark_columns = dataframe_columns[2:]


        exam_groups = {}


        for column in mark_columns:

            parts = str(
                column
            ).split(
                " | "
            )


            if len(parts) < 2:
                continue


            exam_name = parts[0].strip()


            if exam_name not in exam_groups:

                exam_groups[
                    exam_name
                ] = []


            exam_groups[
                exam_name
            ].append(
                column
            )


        # --------------------------------------------------
        # EXAM ORDER
        #
        # QUARTER -> HALF -> ANNUAL
        #
        # Other exams will come afterwards.
        # --------------------------------------------------

        preferred_order = [

            "QUARTER",

            "QUARTERLY",

            "HALF",

            "HALF YEARLY",

            "HALF-YEARLY",

            "HALFYEARLY",

            "ANNUAL",

            "YEARLY",

        ]


        def exam_sort_key(
            exam_name
        ):

            name = str(
                exam_name
            ).strip().upper()


            for index, value in enumerate(
                preferred_order
            ):

                if name == value:

                    return (
                        index,
                        name,
                    )


            # Unknown exams after standard exams

            return (
                100,
                name,
            )


        sorted_exam_names = sorted(

            exam_groups.keys(),

            key=exam_sort_key,

        )


        # --------------------------------------------------
        # CREATE HEADERS
        # --------------------------------------------------

        current_col = 3


        for exam_name in sorted_exam_names:

            exam_columns = exam_groups[
                exam_name
            ]


            exam_start = current_col


            # --------------------------------------------------
            # SUBJECT GROUPS
            # --------------------------------------------------

            subject_groups = {}


            for column in exam_columns:

                parts = str(
                    column
                ).split(
                    " | "
                )


                # --------------------------------------------------
                # EXAM TOTAL
                # Example:
                # QUARTER | TOTAL
                # --------------------------------------------------

                if len(parts) == 2:

                    continue


                if len(parts) < 3:

                    continue


                subject_name = parts[1].strip()


                if subject_name not in subject_groups:

                    subject_groups[
                        subject_name
                    ] = []


                subject_groups[
                    subject_name
                ].append(
                    column
                )


            # --------------------------------------------------
            # SUBJECT HEADINGS
            # --------------------------------------------------

            for subject_name, subject_columns in subject_groups.items():

                subject_start = current_col


                # --------------------------------------------------
                # COMPONENTS
                # --------------------------------------------------

                for column in subject_columns:

                    parts = str(
                        column
                    ).split(
                        " | "
                    )


                    component = parts[-1].strip()


                    ws.cell(

                        row=3,

                        column=current_col,

                        value=component,

                    )


                    current_col += 1


                subject_end = (
                    current_col - 1
                )


                # --------------------------------------------------
                # SUBJECT NAME
                # --------------------------------------------------

                if subject_end >= subject_start:

                    ws.cell(

                        row=2,

                        column=subject_start,

                        value=subject_name,

                    )


                    if subject_end > subject_start:

                        ws.merge_cells(

                            start_row=2,

                            start_column=subject_start,

                            end_row=2,

                            end_column=subject_end,

                        )


            # --------------------------------------------------
            # EXAM TOTAL
            #
            # QUARTER + HALF:
            #     SHOW TOTAL
            #
            # ANNUAL:
            #     DO NOT SHOW TOTAL
            # --------------------------------------------------

            exam_upper = str(
                exam_name
            ).strip().upper()


            is_annual = (

                exam_upper == "ANNUAL"

                or
                exam_upper == "YEARLY"

            )


            total_column = None


            for column in exam_columns:

                parts = str(
                    column
                ).split(
                    " | "
                )


                if len(parts) == 2:

                    if (
                        parts[1]
                        .strip()
                        .upper()
                        == "TOTAL"
                    ):

                        total_column = column

                        break


            # --------------------------------------------------
            # ADD EXAM TOTAL ONLY IF NOT ANNUAL
            # --------------------------------------------------

            if (
                total_column is not None
                and not is_annual
            ):

                ws.cell(

                    row=2,

                    column=current_col,

                    value="TOTAL",

                )


                ws.cell(

                    row=3,

                    column=current_col,

                    value="",

                )


                current_col += 1


            # --------------------------------------------------
            # EXAM HEADING
            # --------------------------------------------------

            exam_end = (
                current_col - 1
            )


            if exam_end >= exam_start:

                ws.cell(

                    row=1,

                    column=exam_start,

                    value=exam_name,

                )


                if exam_end > exam_start:

                    ws.merge_cells(

                        start_row=1,

                        start_column=exam_start,

                        end_row=1,

                        end_column=exam_end,

                    )


        # --------------------------------------------------
        # HEADER STYLE
        # --------------------------------------------------

        for row in range(
            1,
            4
        ):

            for col in range(
                1,
                current_col
            ):

                cell = ws.cell(

                    row=row,

                    column=col,

                )


                cell.font = Font(
                    bold=True
                )


                cell.alignment = Alignment(

                    horizontal="center",

                    vertical="center",

                    wrap_text=True,

                )


                cell.border = thin_border


        # --------------------------------------------------
        # WRITE STUDENT DATA
        #
        # IMPORTANT:
        # DataFrame already contains ONE ROW
        # PER STUDENT.
        #
        # Annual TOTAL is skipped while writing
        # because its Excel column does not exist.
        # --------------------------------------------------

        data_start_row = 4


        # --------------------------------------------------
        # CREATE EXCEL COLUMN MAP
        # --------------------------------------------------

        excel_columns = [

            "Roll No",

            "Student Name",

        ]


        current_excel_col = 3


        for exam_name in sorted_exam_names:

            exam_columns = exam_groups[
                exam_name
            ]


            subject_groups = {}


            for column in exam_columns:

                parts = str(
                    column
                ).split(
                    " | "
                )


                if len(parts) < 3:

                    continue


                subject_name = parts[1].strip()


                if subject_name not in subject_groups:

                    subject_groups[
                        subject_name
                    ] = []


                subject_groups[
                    subject_name
                ].append(
                    column
                )


            for subject_name, subject_columns in subject_groups.items():

                for column in subject_columns:

                    excel_columns.append(
                        column
                    )

                    current_excel_col += 1


            exam_upper = str(
                exam_name
            ).strip().upper()


            is_annual = (

                exam_upper == "ANNUAL"

                or
                exam_upper == "YEARLY"

            )


            # --------------------------------------------------
            # EXAM TOTAL ONLY FOR NON-ANNUAL
            # --------------------------------------------------

            if not is_annual:

                for column in exam_columns:

                    parts = str(
                        column
                    ).split(
                        " | "
                    )


                    if len(parts) == 2:

                        if (
                            parts[1]
                            .strip()
                            .upper()
                            == "TOTAL"
                        ):

                            excel_columns.append(
                                column
                            )

                            current_excel_col += 1

                            break


        # --------------------------------------------------
        # WRITE DATA
        # --------------------------------------------------

        for row_index, row_data in enumerate(

            dataframe.itertuples(

                index=False,

                name=None,

            ),

            start=data_start_row,

        ):

            data_dict = dict(
                zip(
                    dataframe_columns,
                    row_data
                )
            )


            for col_index, column_name in enumerate(

                excel_columns,

                start=1,

            ):

                value = data_dict.get(
                    column_name,
                    ""
                )


                ws.cell(

                    row=row_index,

                    column=col_index,

                    value=value,

                )


        # --------------------------------------------------
        # BORDER + ALIGNMENT
        # --------------------------------------------------

        max_row = ws.max_row

        max_column = ws.max_column


        for row in ws.iter_rows(

            min_row=1,

            max_row=max_row,

            min_col=1,

            max_col=max_column,

        ):

            for cell in row:

                cell.border = thin_border


                cell.alignment = Alignment(

                    horizontal="center",

                    vertical="center",

                    wrap_text=True,

                )


        # --------------------------------------------------
        # STUDENT NAME LEFT ALIGN
        # --------------------------------------------------

        for row in range(

            4,

            max_row + 1

        ):

            ws.cell(

                row=row,

                column=2,

            ).alignment = Alignment(

                horizontal="left",

                vertical="center",

            )


        # --------------------------------------------------
        # COLUMN WIDTH
        # --------------------------------------------------

        ws.column_dimensions[
            "A"
        ].width = 12


        ws.column_dimensions[
            "B"
        ].width = 24


        for col in range(

            3,

            max_column + 1

        ):

            ws.column_dimensions[
                get_column_letter(col)
            ].width = 11


        # --------------------------------------------------
        # ROW HEIGHT
        # --------------------------------------------------

        ws.row_dimensions[
            1
        ].height = 28


        ws.row_dimensions[
            2
        ].height = 30


        ws.row_dimensions[
            3
        ].height = 25


        # --------------------------------------------------
        # FREEZE PANES
        # --------------------------------------------------

        ws.freeze_panes = "C4"


        # --------------------------------------------------
        # SAVE
        # --------------------------------------------------

        wb.save(
            temp_path
        )


        # --------------------------------------------------
        # CHECK FILE
        # --------------------------------------------------

        if os.path.exists(
            temp_path
        ):

            return temp_path


        return None


    except Exception as e:

        print(
            "Consolidated Excel creation error:",
            e
        )


        try:

            if os.path.exists(
                temp_path
            ):

                os.remove(
                    temp_path
                )

        except Exception:

            pass


        return None



# ==========================================================
# LOGIN
# ==========================================================


def login(
    username,
    password,
):

    username = clean_name(
        username
    ).lower()

    password = str(
        password or ""
    )


    if not username or not password:

        return (
            "❌ Enter username and password.",
            gr.update(
                visible=False
            ),
            gr.update(
                visible=False
            ),
            gr.update(
                visible=False
            ),
        )


    db = SessionLocal()

    try:

        teacher = (
            db.query(Teacher)
            .filter(
                Teacher.username == username,
                Teacher.active == True,
            )
            .first()
        )


        if not teacher:

            return (
                "❌ Invalid username or password.",
                gr.update(
                    visible=False
                ),
                gr.update(
                    visible=False
                ),
                gr.update(
                    visible=False
                ),
            )


        if not verify_password(
            password,
            teacher.password_hash,
        ):

            return (
                "❌ Invalid username or password.",
                gr.update(
                    visible=False
                ),
                gr.update(
                    visible=False
                ),
                gr.update(
                    visible=False
                ),
            )


        is_admin = (
            teacher.role == "admin"
            or teacher.username == "admin"
        )


        if is_admin:

            message = (
                f"✅ Welcome, {teacher.name} "
                f"(Admin)"
            )

        else:

            message = (
                f"✅ Welcome, {teacher.name} "
                f"(Teacher)"
            )


        return (
            message,

            gr.update(
                visible=True
            ),

            gr.update(
                visible=is_admin
            ),

            gr.update(
                visible=is_admin
            ),
        )


    finally:

        db.close()


# ==========================================================
# CSS
# ==========================================================


css = """

.gradio-container {
    max-width: 1450px !important;
}

.mobile-mark-input input {
    font-size: 22px !important;
    min-height: 52px !important;
    text-align: center !important;
    -webkit-appearance: none;
    appearance: none;
}

@media (max-width: 768px) {
    .gradio-container {
        padding-left: 10px !important;
        padding-right: 10px !important;
    }

    .mobile-mark-input input {
        font-size: 24px !important;
        min-height: 58px !important;
    }
}

"""


# ==========================================================
# GRADIO APPLICATION
# ==========================================================


with gr.Blocks(
    title="School Mark Entry",
    css=css,
) as demo:


    # ======================================================
    # LOGIN
    # ======================================================

    gr.Markdown(
        "# 🏫 School Mark Entry System"
    )

    gr.Markdown(
        "### 🔐 Login"
    )


    with gr.Row():

        username = gr.Textbox(
            label="Username",
            placeholder="Enter username",
        )

        password = gr.Textbox(
            label="Password",
            type="password",
            placeholder="Enter password",
        )

        login_button = gr.Button(
            "🔐 Login",
            variant="primary",
        )


    login_message = gr.Markdown()


    # ======================================================
    # APPLICATION
    # ======================================================

    with gr.Column(
        visible=False
    ) as application:


        with gr.Tabs():


            # ==================================================
            # MASTER DATA - ADMIN ONLY
            # ==================================================

            with gr.Tab(
                "⚙️ Master Data",
                visible=True,
            ) as master_data_tab:


                gr.Markdown(
                    "## ⚙️ Master Data Management"
                )


                # ==================================================
                # TEACHER MANAGEMENT
                # ==================================================

                gr.Markdown(
                    "### 👨‍🏫 Teacher Management"
                )


                with gr.Row():

                    teacher_username_input = gr.Textbox(
                        label="Teacher Username",
                        placeholder="Example: ramesh",
                    )

                    teacher_name_input = gr.Textbox(
                        label="Teacher Name",
                        placeholder="Example: Ramesh",
                    )

                    teacher_password_input = gr.Textbox(
                        label="Teacher Password",
                        type="password",
                        placeholder="Minimum 6 characters",
                    )


                with gr.Row():

                    add_teacher_button = gr.Button(
                        "➕ Add Teacher",
                        variant="primary",
                    )

                    delete_teacher_select = gr.Dropdown(
                        choices=get_teacher_delete_choices(),
                        label="Select Teacher to Delete",
                    )

                    delete_teacher_button = gr.Button(
                        "🗑️ Delete Teacher",
                        variant="stop",
                    )


                teacher_message = gr.Markdown()


                teacher_table = gr.Dataframe(

                    headers=[
                        "ID",
                        "Username",
                        "Name",
                        "Role",
                        "Status",
                    ],

                    value=get_teacher_list(),

                    interactive=False,

                )


                gr.Markdown(
                    """
**Teacher Login:**  
Admin creates each teacher's username and password here.

Teacher accounts can use **Mark Entry, View Marks and Consolidated Report**, but cannot access Admin Master Data or Student Management.
"""
                )


                # ==================================================
                # ACADEMIC YEAR
                # ==================================================

                gr.Markdown(
                    "### 📅 Academic Year"
                )


                with gr.Row():

                    academic_year_name = gr.Textbox(
                        label="Academic Year",
                        placeholder="Example: 2027-28",
                    )

                    add_year_button = gr.Button(
                        "➕ Add Year",
                        variant="primary",
                    )


                with gr.Row():

                    delete_year_select = gr.Dropdown(
                        choices=get_years(),
                        label="Select Year to Delete",
                    )

                    delete_year_button = gr.Button(
                        "🗑️ Delete Year",
                        variant="stop",
                    )


                year_message = gr.Markdown()


                year_table = gr.Dataframe(

                    headers=[
                        "ID",
                        "Academic Year",
                        "Status",
                    ],

                    value=get_year_list(),

                    interactive=False,

                )


                # ==================================================
                # CLASS
                # ==================================================

                gr.Markdown(
                    "### 🏫 Class"
                )


                with gr.Row():

                    class_name_input = gr.Textbox(
                        label="Class Name",
                        placeholder="Example: 12-C",
                    )

                    add_class_button = gr.Button(
                        "➕ Add Class",
                        variant="primary",
                    )


                with gr.Row():

                    delete_class_select = gr.Dropdown(
                        choices=get_classes(),
                        label="Select Class to Delete",
                    )

                    delete_class_button = gr.Button(
                        "🗑️ Delete Class",
                        variant="stop",
                    )


                class_message = gr.Markdown()


                class_table = gr.Dataframe(

                    headers=[
                        "ID",
                        "Class",
                    ],

                    value=get_class_list(),

                    interactive=False,

                )


                # ==================================================
                # SUBJECT
                # ==================================================

                gr.Markdown(
                    "### 📚 Subject"
                )


                with gr.Row():

                    subject_name_input = gr.Textbox(
                        label="Subject Name",
                        placeholder="Example: Chemistry",
                    )

                    subject_code_input = gr.Textbox(
                        label="Subject Code",
                        placeholder="Example: CHE",
                    )


                with gr.Row():

                    subject_theory = gr.Checkbox(
                        label="Theory",
                        value=True,
                    )

                    subject_practical = gr.Checkbox(
                        label="Practical",
                        value=False,
                    )

                    subject_internal = gr.Checkbox(
                        label="Internal",
                        value=True,
                    )


                with gr.Row():

                    subject_theory_max = gr.Number(
                        label="Theory Maximum",
                        value=80,
                        precision=0,
                    )

                    subject_practical_max = gr.Number(
                        label="Practical Maximum",
                        value=0,
                        precision=0,
                    )

                    subject_internal_max = gr.Number(
                        label="Internal Maximum",
                        value=20,
                        precision=0,
                    )


                with gr.Row():

                    add_subject_button = gr.Button(
                        "➕ Add Subject",
                        variant="primary",
                    )

                    delete_subject_select = gr.Dropdown(
                        choices=get_subject_choices(),
                        label="Select Subject to Delete",
                    )

                    delete_subject_button = gr.Button(
                        "🗑️ Delete Subject",
                        variant="stop",
                    )


                subject_message = gr.Markdown()


                subject_table = gr.Dataframe(

                    headers=[
                        "ID",
                        "Subject",
                        "Code",
                        "Components",
                        "Status",
                    ],

                    value=get_subject_list(),

                    interactive=False,

                )


                # ==================================================
                # EXAM
                # ==================================================

                gr.Markdown(
                    "### 📝 Exam"
                )


                with gr.Row():

                    exam_name_input = gr.Textbox(
                        label="Exam Name",
                        placeholder="Example: Monthly Test",
                    )

                    add_exam_button = gr.Button(
                        "➕ Add Exam",
                        variant="primary",
                    )


                with gr.Row():

                    delete_exam_select = gr.Dropdown(
                        choices=get_exams(),
                        label="Select Exam to Delete",
                    )

                    delete_exam_button = gr.Button(
                        "🗑️ Delete Exam",
                        variant="stop",
                    )


                exam_message = gr.Markdown()


                exam_table = gr.Dataframe(

                    headers=[
                        "ID",
                        "Exam",
                        "Status",
                    ],

                    value=get_exam_list(),

                    interactive=False,

                )


            # ==================================================
            # STUDENT MANAGEMENT - ADMIN ONLY
            # ==================================================

            with gr.Tab(
                "👨‍🎓 Student Management",
                visible=True,
            ) as student_management_tab:


                gr.Markdown(
                    "## 👨‍🎓 Student Management"
                )


                with gr.Row():

                    admission_no = gr.Textbox(
                        label="Admission No",
                        placeholder="Admission Number",
                    )

                    roll_no = gr.Textbox(
                        label="Roll No",
                        placeholder="Roll Number",
                    )

                    student_name = gr.Textbox(
                        label="Student Name",
                        placeholder="Student Name",
                    )

                    student_class = gr.Dropdown(
                        choices=get_classes(),
                        label="Class",
                    )


                with gr.Row():

                    add_student_button = gr.Button(
                        "➕ Add Student",
                        variant="primary",
                    )

                    delete_student_select = gr.Dropdown(
                        choices=get_student_delete_choices(),
                        label="Select Student to Delete",
                    )

                    delete_student_button = gr.Button(
                        "🗑️ Delete Student",
                        variant="stop",
                    )


                student_message = gr.Markdown()


                student_table = gr.Dataframe(

                    headers=[
                        "ID",
                        "Admission No",
                        "Roll No",
                        "Student Name",
                        "Class",
                    ],

                    value=get_student_list(),

                    interactive=False,

                )


            # ==================================================
            # MARK ENTRY - MOBILE FRIENDLY
            # ==================================================

            with gr.Tab(
                "📝 Mark Entry"
            ):

                gr.Markdown(
                    "## 📝 Mark Entry"
                )

                with gr.Row():

                    mark_year = gr.Dropdown(
                        choices=get_years(),
                        label="Academic Year",
                    )

                    mark_class = gr.Dropdown(
                        choices=get_classes(),
                        label="Class",
                    )

                    mark_exam = gr.Dropdown(
                        choices=get_exams(),
                        label="Exam",
                    )

                    mark_subject = gr.Dropdown(
                        choices=[],
                        label="Subject",
                    )

                mark_class.change(
                    get_subjects_for_class,
                    inputs=mark_class,
                    outputs=mark_subject,
                )

                gr.Markdown(
                    "📱 **Mobile Mark Entry:** Tap a mark box and enter the mark using your phone's number keyboard."
                )

                @gr.render(
                    inputs=[
                        mark_year,
                        mark_class,
                        mark_exam,
                        mark_subject,
                    ]
                )
                def render_mobile_mark_entry(
                    year_value,
                    class_value,
                    exam_value,
                    subject_value,
                ):
                    class_id = parse_id(class_value)
                    year_id = parse_id(year_value)
                    exam_id = parse_id(exam_value)
                    subject_id = parse_id(subject_value)

                    if not class_id:
                        gr.Markdown("⬆️ Select a **Class** to begin.")
                        return

                    if not year_id:
                        gr.Markdown("⬆️ Select **Academic Year**.")
                        return

                    if not exam_id:
                        gr.Markdown("⬆️ Select **Exam**.")
                        return

                    if not subject_id:
                        gr.Markdown("⬆️ Select **Subject**.")
                        return

                    db = SessionLocal()

                    try:
                        subject = db.get(Subject, subject_id)

                        if subject is None or not subject.active:
                            gr.Markdown("❌ Subject not found.")
                            return

                        students = (
                            db.query(Student)
                            .filter(
                                Student.class_id == class_id,
                                Student.active == True,
                            )
                            .order_by(
                                Student.roll_no,
                                Student.name,
                            )
                            .all()
                        )

                        if not students:
                            gr.Markdown(
                                "❌ No active students found in this class."
                            )
                            return

                        pattern = []

                        if subject.theory:
                            pattern.append(
                                f"Theory: {subject.theory_max}"
                            )

                        if subject.practical:
                            pattern.append(
                                f"Practical: {subject.practical_max}"
                            )

                        if subject.internal:
                            pattern.append(
                                f"Internal: {subject.internal_max}"
                            )

                        gr.Markdown(
                            "**Mark Pattern:** "
                            + " + ".join(pattern)
                        )

                        gr.Markdown(
                            "### 👨‍🎓 Enter Marks"
                        )

                        mark_inputs = []

                        for student in students:
                            mark = (
                                db.query(Mark)
                                .filter_by(
                                    academic_year_id=year_id,
                                    exam_id=exam_id,
                                    student_id=student.id,
                                    subject_id=subject_id,
                                )
                                .first()
                            )

                            with gr.Group():
                                gr.Markdown(
                                    f"**Roll No: {student.roll_no or '-'} — {student.name}**"
                                )

                                with gr.Row():

                                    if subject.theory:
                                        theory_value = (
                                            mark.theory
                                            if mark
                                            else 0
                                        )
                                        theory_input = gr.Number(
                                            value=theory_value,
                                            label=f"Theory / {subject.theory_max}",
                                            precision=0,
                                            minimum=0,
                                            maximum=subject.theory_max,
                                            step=1,
                                            interactive=True,
                                            elem_classes=[
                                                "mobile-mark-input"
                                            ],
                                        )
                                        mark_inputs.append(
                                            theory_input
                                        )

                                    if subject.practical:
                                        practical_value = (
                                            mark.practical
                                            if mark
                                            else 0
                                        )
                                        practical_input = gr.Number(
                                            value=practical_value,
                                            label=f"Practical / {subject.practical_max}",
                                            precision=0,
                                            minimum=0,
                                            maximum=subject.practical_max,
                                            step=1,
                                            interactive=True,
                                            elem_classes=[
                                                "mobile-mark-input"
                                            ],
                                        )
                                        mark_inputs.append(
                                            practical_input
                                        )

                                    if subject.internal:
                                        internal_value = (
                                            mark.internal
                                            if mark
                                            else 0
                                        )
                                        internal_input = gr.Number(
                                            value=internal_value,
                                            label=f"Internal / {subject.internal_max}",
                                            precision=0,
                                            minimum=0,
                                            maximum=subject.internal_max,
                                            step=1,
                                            interactive=True,
                                            elem_classes=[
                                                "mobile-mark-input"
                                            ],
                                        )
                                        mark_inputs.append(
                                            internal_input
                                        )

                        save_button = gr.Button(
                            "💾 Save Marks",
                            variant="primary",
                        )

                        save_message = gr.Markdown()

                        save_button.click(
                            save_mobile_marks,
                            inputs=[
                                mark_class,
                                mark_year,
                                mark_exam,
                                mark_subject,
                                *mark_inputs,
                            ],
                            outputs=save_message,
                        )

                    finally:
                        db.close()

            # ==================================================
            # VIEW MARKS
            # ==================================================

            with gr.Tab(
                "👁️ View Marks"
            ):

                gr.Markdown(
                    "## 👁️ Subject-wise Mark View"
                )


                with gr.Row():

                    view_year = gr.Dropdown(
                        choices=get_years(),
                        label="Academic Year",
                    )

                    view_class = gr.Dropdown(
                        choices=get_classes(),
                        label="Class",
                    )

                    view_exam = gr.Dropdown(
                        choices=get_exams(),
                        label="Exam",
                    )

                    view_subject = gr.Dropdown(
                        choices=[],
                        label="Subject",
                    )


                view_class.change(

                    get_subjects_for_class,

                    inputs=view_class,

                    outputs=view_subject,

                )


                view_button = gr.Button(
                    "👁️ View Marks"
                )


                view_pattern = gr.Markdown()


                view_table = gr.Dataframe(
                    interactive=False,
                    wrap=True,
                )


                view_button.click(

                    load_marks,

                    inputs=[

                        view_class,

                        view_year,

                        view_exam,

                        view_subject,

                    ],

                    outputs=[

                        view_table,

                        view_pattern,

                    ],

                )


                # --------------------------------------------------
                # VIEW MARKS EXCEL
                # --------------------------------------------------

                view_excel_button = gr.DownloadButton(
                    "📥 Download View Marks Excel",
                    variant="secondary",
                )


                view_excel_button.click(

                    download_view_marks_excel,

                    inputs=view_table,

                    outputs=view_excel_button,

                )


            # ==================================================
            # CONSOLIDATED REPORT
            # ==================================================

            with gr.Tab(
                "📊 Mark List / Print"
            ):

                gr.Markdown(
                    "## 📊 Consolidated Mark List"
                )


                gr.Markdown(
                    """
Select **Academic Year + Class + Exam**.

The report will show:

**Roll No → Student → Subject → Exam → Theory → Practical → Internal → Total**
"""
                )


                with gr.Row():

                    list_year = gr.Dropdown(
                        choices=get_years(),
                        label="Academic Year",
                    )

                    list_class = gr.Dropdown(
                        choices=get_classes(),
                        label="Class",
                    )

                    list_exam = gr.Dropdown(
                        choices=get_all_exams(),
                        label="Exam",
                    )


                consolidated_button = gr.Button(

                    "📋 View Consolidated Mark List",

                    variant="primary",

                )


                consolidated_message = gr.Markdown()


                consolidated_table = gr.Dataframe(

                    interactive=False,

                    wrap=True,

                )


                # --------------------------------------------------
                # CONSOLIDATED EXCEL
                # --------------------------------------------------

                consolidated_excel_button = gr.DownloadButton(
                    "📥 Download Consolidated Excel",
                    variant="secondary",
                        interactive=True,
                    )

                consolidated_excel_button.click(
                    fn=download_consolidated_excel,
                    inputs=[consolidated_table],
                    outputs=[consolidated_excel_button],
                )


                gr.Markdown(
                    """
### 🖨️ Print

After the report appears,
use **Ctrl + P** in your browser to print.
"""
                )


                consolidated_button.click(

                    generate_consolidated_report,

                    inputs=[

                        list_year,

                        list_class,

                        list_exam,

                    ],

                    outputs=[

                        consolidated_message,

                        consolidated_table,

                    ],

                )


    # ======================================================
    # TEACHER EVENTS
    # ======================================================

    add_teacher_button.click(

        add_teacher,

        inputs=[

            teacher_username_input,

            teacher_name_input,

            teacher_password_input,

        ],

        outputs=[

            teacher_message,

            teacher_table,

            delete_teacher_select,

        ],

    )


    delete_teacher_button.click(

        delete_teacher,

        inputs=delete_teacher_select,

        outputs=[

            teacher_message,

            teacher_table,

            delete_teacher_select,

        ],

    )


    # ======================================================
    # STUDENT EVENTS
    # ======================================================

    add_student_button.click(

        add_student,

        inputs=[

            admission_no,

            roll_no,

            student_name,

            student_class,

        ],

        outputs=[

            student_message,

            student_table,

            delete_student_select,

        ],

    )


    delete_student_button.click(

        delete_student,

        inputs=delete_student_select,

        outputs=[

            student_message,

            student_table,

            delete_student_select,

        ],

    )


    # ======================================================
    # MASTER DATA EVENTS
    # ======================================================

    add_year_button.click(

        add_academic_year,

        inputs=academic_year_name,

        outputs=[

            year_message,

            year_table,

            delete_year_select,

            mark_year,

            view_year,

            list_year,

        ],

    )


    delete_year_button.click(

        delete_academic_year,

        inputs=delete_year_select,

        outputs=[

            year_message,

            year_table,

            delete_year_select,

            mark_year,

            view_year,

            list_year,

        ],

    )


    add_class_button.click(

        add_class,

        inputs=class_name_input,

        outputs=[

            class_message,

            class_table,

            delete_class_select,

            student_class,

            mark_class,

            view_class,

            list_class,

        ],

    )


    delete_class_button.click(

        delete_class,

        inputs=delete_class_select,

        outputs=[

            class_message,

            class_table,

            delete_class_select,

            student_class,

            mark_class,

            view_class,

            list_class,

        ],

    )


    add_subject_button.click(

        add_subject,

        inputs=[

            subject_name_input,

            subject_code_input,

            subject_theory,

            subject_practical,

            subject_internal,

            subject_theory_max,

            subject_practical_max,

            subject_internal_max,

        ],

        outputs=[

            subject_message,

            subject_table,

            delete_subject_select,

        ],

    )


    delete_subject_button.click(

        delete_subject,

        inputs=delete_subject_select,

        outputs=[

            subject_message,

            subject_table,

            delete_subject_select,

        ],

    )


    add_exam_button.click(

        add_exam,

        inputs=exam_name_input,

        outputs=[

            exam_message,

            exam_table,

            delete_exam_select,

            mark_exam,

            view_exam,

            list_exam,

        ],

    )


    delete_exam_button.click(

        delete_exam,

        inputs=delete_exam_select,

        outputs=[

            exam_message,

            exam_table,

            delete_exam_select,

            mark_exam,

            view_exam,

            list_exam,

        ],

    )


    # ======================================================
    # LOGIN EVENT
    # ======================================================

    login_button.click(

        login,

        inputs=[

            username,

            password,

        ],

        outputs=[

            login_message,

            application,

            master_data_tab,

            student_management_tab,

        ],

    )


    # ======================================================
    # LOGIN INFORMATION
    # ======================================================

    gr.Markdown(
        "**Initial Admin Login:** `admin` / `Admin@123`"
    )


# ==========================================================
# START SERVER
# ==========================================================


if __name__ == "__main__":

    port = int(
        os.getenv(
            "PORT",
            "7860",
        )
    )


    demo.launch(

        server_name="0.0.0.0",

        server_port=port,

    )
